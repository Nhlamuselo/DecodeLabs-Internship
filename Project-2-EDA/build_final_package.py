from pathlib import Path
import math
import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, BarChart, DoughnutChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as PdfTable, TableStyle, PageBreak
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie

INPUT = Path(r"C:\Users\prais\Downloads\Dataset for Data Analytics (1).xlsx")
OUTDIR = Path(r"C:\Users\prais\Documents\Codex\2026-05-20\files-mentioned-by-the-user-data\outputs\data-analytics-p2")
OUTDIR.mkdir(parents=True, exist_ok=True)
WIDESCREEN = (13.333 * inch, 7.5 * inch)

df = pd.read_excel(INPUT)
df["Date"] = pd.to_datetime(df["Date"])
df["Month"] = df["Date"].dt.to_period("M").astype(str)
df["Quarter"] = df["Date"].dt.to_period("Q").astype(str).str.replace("Q", " Q", regex=False)

total_orders = len(df)
total_value = float(df["TotalPrice"].sum())
avg_order = float(df["TotalPrice"].mean())
delivered_value = float(df.loc[df["OrderStatus"].eq("Delivered"), "TotalPrice"].sum())
friction = df[df["OrderStatus"].isin(["Cancelled", "Returned"])]
friction_value = float(friction["TotalPrice"].sum())
q1 = float(df["TotalPrice"].quantile(0.25))
q3 = float(df["TotalPrice"].quantile(0.75))
iqr = q3 - q1
upper_fence = q3 + 1.5 * iqr
outliers = df[df["TotalPrice"] > upper_fence].sort_values("TotalPrice", ascending=False)

def seg(group):
    return (
        df.groupby(group)
        .agg(Orders=("OrderID", "count"), Revenue=("TotalPrice", "sum"), Quantity=("Quantity", "sum"), Avg_Order=("TotalPrice", "mean"))
        .reset_index()
        .sort_values("Revenue", ascending=False)
    )

product = seg("Product")
source = seg("ReferralSource")
payment = seg("PaymentMethod")
status = seg("OrderStatus")
monthly = (
    df.groupby("Month")
    .agg(Orders=("OrderID", "count"), Revenue=("TotalPrice", "sum"), Avg_Order=("TotalPrice", "mean"))
    .reset_index()
    .sort_values("Month")
)
best_month = monthly.sort_values("Revenue", ascending=False).iloc[0]
latest_month = monthly.iloc[-1]
prior_month = monthly.iloc[-2]
latest_mom = (latest_month["Revenue"] / prior_month["Revenue"] - 1) * 100
top_product = product.iloc[0]
top_source = source.iloc[0]

NAVY = "172A3A"
TEAL = "0B6E69"
MINT = "DDEEEA"
CREAM = "F7F4EF"
GOLD = "C9972B"
RED = "B94747"
GRAY = "65717C"
WHITE = "FFFFFF"
LINE = "D6DDD9"

thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def money(v):
    return f"${v:,.0f}"

def pct(v):
    return f"{v:.1f}%"

def header(ws, cell_range, title, subtitle):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=18)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row = cell.row + 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    sub = ws.cell(row, 1)
    sub.value = subtitle
    sub.fill = PatternFill("solid", fgColor=MINT)
    sub.font = Font(color=NAVY, italic=True)
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[cell.row].height = 28

def style_table(ws, start_row, start_col, end_row, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(start_row, c)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

def write_frame(ws, df_frame, row, col):
    for j, name in enumerate(df_frame.columns, col):
        ws.cell(row, j).value = name.replace("_", " ")
    for i, (_, rec) in enumerate(df_frame.iterrows(), row + 1):
        for j, val in enumerate(rec, col):
            ws.cell(i, j).value = val.item() if hasattr(val, "item") else val
    style_table(ws, row, col, row + len(df_frame), col + len(df_frame.columns) - 1)
    return row + len(df_frame), col + len(df_frame.columns) - 1

wb = Workbook()
ws = wb.active
ws.title = "Executive Dashboard"
ws.sheet_view.showGridLines = False
header(ws, "A1:N1", "Ecommerce EDA Dashboard", "Project 2: descriptive statistics, trends, outliers, and key business observations")

kpis = [
    ("Total Order Value", total_value, "$#,##0"),
    ("Orders", total_orders, "#,##0"),
    ("Average Order Value", avg_order, "$#,##0"),
    ("Delivered Value", delivered_value, "$#,##0"),
    ("Returned + Cancelled Value", friction_value, "$#,##0"),
    ("Outlier Orders", len(outliers), "#,##0"),
]
for i, (label, value, fmt) in enumerate(kpis):
    col = 1 + i * 2
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
    ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
    ws.cell(4, col).value = label
    ws.cell(5, col).value = value
    ws.cell(4, col).font = Font(color=GRAY, bold=True, size=9)
    ws.cell(5, col).font = Font(color=RED if i == 4 else TEAL, bold=True, size=15)
    ws.cell(5, col).number_format = fmt
    for r in (4, 5):
        for c in range(col, col + 2):
            ws.cell(r, c).fill = PatternFill("solid", fgColor="F6E5E1" if i == 4 else WHITE)
            ws.cell(r, c).border = border
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A7:N7")
ws["A7"] = "Executive readout"
ws["A7"].fill = PatternFill("solid", fgColor=TEAL)
ws["A7"].font = Font(color=WHITE, bold=True)
ws.merge_cells("A8:N12")
ws["A8"] = (
    f"The dataset records {total_orders:,} ecommerce orders worth {money(total_value)} between "
    f"{monthly.iloc[0]['Month']} and {latest_month['Month']}. {top_product['Product']} leads product value at "
    f"{money(top_product['Revenue'])}, while {top_source['ReferralSource']} is the strongest recorded acquisition source. "
    f"The largest opportunity is operational: returned and cancelled orders represent {money(friction_value)}, a material "
    f"amount of order value that investors would expect management to reduce. The final observed month, {latest_month['Month']}, "
    f"improved {pct(latest_mom)} month over month."
)
ws["A8"].fill = PatternFill("solid", fgColor=CREAM)
ws["A8"].font = Font(color=NAVY, size=11)
ws["A8"].alignment = Alignment(wrap_text=True, vertical="top")
for r in range(8, 13):
    ws.row_dimensions[r].height = 23

chart_data = monthly[["Month", "Revenue"]]
write_frame(ws, chart_data, 15, 1)
for r in range(16, 46):
    ws.cell(r, 2).number_format = "$#,##0"
line = LineChart()
line.title = "Monthly Order Value Trend"
line.y_axis.title = "Order value"
line.x_axis.title = "Month"
line.height = 8
line.width = 20
line.add_data(Reference(ws, min_col=2, min_row=15, max_row=45), titles_from_data=True)
line.set_categories(Reference(ws, min_col=1, min_row=16, max_row=45))
line.style = 13
ws.add_chart(line, "D15")

write_frame(ws, product[["Product", "Revenue"]], 49, 1)
for r in range(50, 57):
    ws.cell(r, 2).number_format = "$#,##0"
bar = BarChart()
bar.title = "Revenue by Product"
bar.y_axis.title = "Revenue"
bar.x_axis.title = "Product"
bar.height = 7
bar.width = 13
bar.add_data(Reference(ws, min_col=2, min_row=49, max_row=56), titles_from_data=True)
bar.set_categories(Reference(ws, min_col=1, min_row=50, max_row=56))
ws.add_chart(bar, "D36")

write_frame(ws, status[["OrderStatus", "Revenue"]], 36, 11)
for r in range(37, 42):
    ws.cell(r, 12).number_format = "$#,##0"
pie = DoughnutChart()
pie.title = "Order Value by Status"
pie.add_data(Reference(ws, min_col=12, min_row=36, max_row=41), titles_from_data=True)
pie.set_categories(Reference(ws, min_col=11, min_row=37, max_row=41))
pie.height = 7
pie.width = 8
ws.add_chart(pie, "K43")

insights = [
    ["Investor angle", "What it means", "Evidence", "Action"],
    ["Product concentration", "Revenue is balanced enough to scale without a single-product dependency.", f"{top_product['Product']} is #1, but top product revenue is tightly clustered.", "Keep broad assortment, then test bundles around top products."],
    ["Channel quality", "Instagram and Email provide the strongest recorded order value.", f"{top_source['ReferralSource']}: {money(top_source['Revenue'])} order value.", "Shift experiment budget toward high-AOV sources."],
    ["Operational drag", "Cancelled/returned value is the clearest margin protection lever.", f"{money(friction_value)} in returned/cancelled order value.", "Audit product descriptions, delivery promise, and post-purchase support."],
    ["High-value outliers", "Largest orders are real expansion signals, not automatic errors.", f"{len(outliers)} orders above {money(upper_fence)}.", "Create VIP retention workflows for large baskets."],
    ["Demand recency", "The final observed month rebounded meaningfully.", f"{latest_month['Month']}: {money(latest_month['Revenue'])}, {pct(latest_mom)} MoM.", "Use this as the leading slide in investor discussions."],
]
for i, row in enumerate(insights, 60):
    for j, val in enumerate(row, 1):
        ws.cell(i, j).value = val
style_table(ws, 60, 1, 65, 4)
for r in range(61, 66):
    ws.row_dimensions[r].height = 45

stats_ws = wb.create_sheet("EDA Statistics")
stats_ws.sheet_view.showGridLines = False
header(stats_ws, "A1:H1", "Exploratory Statistics", "Mean, median, count, spread, and distribution checks")
metrics = []
for col in ["Quantity", "UnitPrice", "ItemsInCart", "TotalPrice"]:
    s = df[col]
    metrics.append([col, s.count(), s.mean(), s.median(), s.min(), s.max(), s.std()])
for i, row in enumerate([["Metric", "Count", "Mean", "Median", "Min", "Max", "Std Dev"], *metrics], 4):
    for j, val in enumerate(row, 1):
        stats_ws.cell(i, j).value = val
style_table(stats_ws, 4, 1, 8, 7)
for row in range(5, 9):
    for col in range(3, 8):
        stats_ws.cell(row, col).number_format = "$#,##0.00" if stats_ws.cell(row, 1).value in ["UnitPrice", "TotalPrice"] else "0.00"
outlier_table = [
    ["Outlier method", "Value"],
    ["Q1 total price", q1],
    ["Q3 total price", q3],
    ["IQR", iqr],
    ["Upper fence", upper_fence],
    ["Outlier count", len(outliers)],
    ["Largest order", df["TotalPrice"].max()],
    ["Median order", df["TotalPrice"].median()],
]
for i, row in enumerate(outlier_table, 11):
    for j, val in enumerate(row, 1):
        stats_ws.cell(i, j).value = val
style_table(stats_ws, 11, 1, 18, 2)
for r in range(12, 19):
    stats_ws.cell(r, 2).number_format = "$#,##0.00" if r != 16 else "0"
corr = df[["Quantity", "UnitPrice", "ItemsInCart", "TotalPrice"]].corr().round(3)
for i, row in enumerate([["Correlation", *corr.columns], *[[idx, *corr.loc[idx].tolist()] for idx in corr.index]], 11):
    for j, val in enumerate(row, 4):
        stats_ws.cell(i, j).value = val
style_table(stats_ws, 11, 4, 15, 8)
stats_ws.conditional_formatting.add("E12:H15", ColorScaleRule(start_type="min", start_color="F6E5E1", mid_type="percentile", mid_value=50, mid_color="FFFFFF", end_type="max", end_color="DDEEEA"))

seg_ws = wb.create_sheet("Segments")
seg_ws.sheet_view.showGridLines = False
header(seg_ws, "A1:N1", "Segment Analysis", "Product, source, payment method, order status, and monthly views")
write_frame(seg_ws, product, 4, 1)
write_frame(seg_ws, source, 4, 7)
write_frame(seg_ws, payment, 15, 1)
write_frame(seg_ws, status, 15, 7)
write_frame(seg_ws, monthly, 26, 1)
for row in seg_ws.iter_rows():
    for cell in row:
        if str(cell.value).replace(" ", "_") in ["Revenue", "Avg_Order"]:
            pass
for col in [3, 5, 9, 11]:
    for r in range(5, 60):
        seg_ws.cell(r, col).number_format = "$#,##0"

out_ws = wb.create_sheet("Outliers")
out_ws.sheet_view.showGridLines = False
header(out_ws, "A1:J1", "High-Value Outliers", "Orders above the IQR upper fence; review as high-value signals, not automatic errors")
out_cols = ["OrderID", "Date", "Product", "Quantity", "UnitPrice", "TotalPrice", "OrderStatus", "ReferralSource"]
out_table = outliers[out_cols].copy()
out_table["Comment"] = "Large basket: validate customer experience and retention follow-up"
write_frame(out_ws, out_table, 4, 1)
for r in range(5, 5 + len(out_table)):
    out_ws.cell(r, 2).number_format = "yyyy-mm-dd"
    out_ws.cell(r, 5).number_format = "$#,##0.00"
    out_ws.cell(r, 6).number_format = "$#,##0.00"

raw_ws = wb.create_sheet("Raw Data")
raw_ws.sheet_view.showGridLines = False
for j, col in enumerate(df.columns, 1):
    raw_ws.cell(1, j).value = col
for i, row in enumerate(df.itertuples(index=False), 2):
    for j, val in enumerate(row, 1):
        raw_ws.cell(i, j).value = val.to_pydatetime() if hasattr(val, "to_pydatetime") else val
style_table(raw_ws, 1, 1, len(df) + 1, len(df.columns))
tab = Table(displayName="OrdersTable", ref=f"A1:{get_column_letter(len(df.columns))}{len(df)+1}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
raw_ws.add_table(tab)
raw_ws.freeze_panes = "A2"
for r in range(2, len(df) + 2):
    raw_ws.cell(r, 2).number_format = "yyyy-mm-dd"
    raw_ws.cell(r, 6).number_format = "$#,##0.00"
    raw_ws.cell(r, 14).number_format = "$#,##0.00"

method_ws = wb.create_sheet("Methodology")
method_ws.sheet_view.showGridLines = False
header(method_ws, "A1:H1", "Methodology and Data Notes", "Audit trail for the EDA work")
notes = [
    ["Area", "Treatment"],
    ["Source file", "Dataset for Data Analytics (1).xlsx"],
    ["Rows analyzed", total_orders],
    ["Date coverage", f"{monthly.iloc[0]['Month']} to {latest_month['Month']}"],
    ["Core statistics", "Count, mean, median, min, max, standard deviation, quartiles, and IQR outlier fence."],
    ["Trend analysis", "Monthly order value and order count were grouped from transaction dates."],
    ["Segment analysis", "Revenue and order count were compared by product, referral source, payment method, and status."],
    ["Outlier definition", f"TotalPrice greater than Q3 + 1.5 x IQR = {money(upper_fence)}."],
    ["Business interpretation", "Insights focus on growth signals, channel performance, product mix, and operational leakage."],
    ["Important note", "Order status values include delivered, shipped, pending, cancelled, and returned; recorded order value is treated as demand/order value, not confirmed recognized revenue."],
]
for i, row in enumerate(notes, 4):
    for j, val in enumerate(row, 1):
        method_ws.cell(i, j).value = val
style_table(method_ws, 4, 1, 13, 2)

for sheet in wb.worksheets:
    for col in range(1, min(sheet.max_column, 16) + 1):
        letter = get_column_letter(col)
        sheet.column_dimensions[letter].width = 16
    if sheet.title in ["Executive Dashboard"]:
        for col in range(1, 15):
            sheet.column_dimensions[get_column_letter(col)].width = 13
        sheet.column_dimensions["D"].width = 16
    if sheet.title == "Raw Data":
        sheet.column_dimensions["G"].width = 24
    if sheet.title in ["Outliers", "Methodology"]:
        sheet.column_dimensions["I"].width = 42
        sheet.column_dimensions["B"].width = 48 if sheet.title == "Methodology" else 16
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = cell.font.copy(name="Aptos")
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")

xlsx_path = OUTDIR / "Data Analytics P2 - Ecommerce EDA Dashboard.xlsx"
wb.save(xlsx_path)

styles = getSampleStyleSheet()
styles.add(ParagraphStyle(name="SlideTitle", parent=styles["Title"], fontSize=28, leading=32, textColor=colors.HexColor("#" + NAVY), spaceAfter=12))
styles.add(ParagraphStyle(name="Subtle", parent=styles["BodyText"], fontSize=12, leading=16, textColor=colors.HexColor("#" + GRAY)))
styles.add(ParagraphStyle(name="Big", parent=styles["Heading1"], fontSize=22, leading=26, textColor=colors.HexColor("#" + TEAL)))

pdf_path = OUTDIR / "Data Analytics P2 - Investor Brief.pdf"
doc = SimpleDocTemplate(str(pdf_path), pagesize=WIDESCREEN, rightMargin=0.45*inch, leftMargin=0.45*inch, topMargin=0.35*inch, bottomMargin=0.35*inch)
story = []

def add_title(title, subtitle=None):
    story.append(Paragraph(title, styles["SlideTitle"]))
    if subtitle:
        story.append(Paragraph(subtitle, styles["Subtle"]))
    story.append(Spacer(1, 0.18 * inch))

add_title("Ecommerce Order Analytics", "Exploratory data analysis for Project 2 | Built from 1,200 transaction records")
kpi_table = PdfTable([
    ["Total order value", "Orders", "Avg order", "Operational leakage"],
    [money(total_value), f"{total_orders:,}", money(avg_order), money(friction_value)],
], colWidths=[2.5*inch]*4)
kpi_table.setStyle(TableStyle([
    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#" + NAVY)),
    ("TEXTCOLOR", (0,0), (-1,0), colors.white),
    ("BACKGROUND", (0,1), (-1,1), colors.HexColor("#" + CREAM)),
    ("TEXTCOLOR", (0,1), (-1,1), colors.HexColor("#" + TEAL)),
    ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
    ("FONTSIZE", (0,1), (-1,1), 20),
    ("ALIGN", (0,0), (-1,-1), "CENTER"),
    ("BOX", (0,0), (-1,-1), 0.5, colors.HexColor("#" + LINE)),
    ("INNERGRID", (0,0), (-1,-1), 0.25, colors.HexColor("#" + LINE)),
]))
story.append(kpi_table)
story.append(Spacer(1, 0.3 * inch))
story.append(Paragraph(f"The story is not just that the store has demand. The stronger investor point is that demand is diversified across products and channels, while the largest upside sits in reducing returned and cancelled order value.", styles["Big"]))
story.append(PageBreak())

add_title("Where growth is coming from", f"{top_product['Product']} leads product value and {top_source['ReferralSource']} is the strongest source.")
d = Drawing(760, 250)
bc = VerticalBarChart()
bc.x = 55
bc.y = 35
bc.height = 180
bc.width = 650
bc.data = [list(product["Revenue"])]
bc.categoryAxis.categoryNames = list(product["Product"])
bc.valueAxis.valueMin = 0
bc.valueAxis.valueMax = math.ceil(product["Revenue"].max() / 50000) * 50000
bc.bars[0].fillColor = colors.HexColor("#" + TEAL)
d.add(bc)
story.append(d)
story.append(Paragraph("The product mix is healthy: no single product carries the whole business. That gives management room to optimize bundles, campaigns, and service quality without betting everything on one SKU.", styles["Subtle"]))
story.append(PageBreak())

add_title("Trend signal", f"Best month: {best_month['Month']} at {money(best_month['Revenue'])}; latest month improved {pct(latest_mom)} month over month.")
d = Drawing(760, 250)
lc = HorizontalLineChart()
lc.x = 55
lc.y = 35
lc.height = 180
lc.width = 650
lc.data = [list(monthly["Revenue"])]
lc.categoryAxis.categoryNames = list(monthly["Month"])
lc.categoryAxis.labels.angle = 45
lc.lines[0].strokeColor = colors.HexColor("#" + GOLD)
lc.lines[0].strokeWidth = 2
d.add(lc)
story.append(d)
story.append(Paragraph("The monthly series is uneven, but the final observed month shows renewed momentum. That is the right place to frame the next commercial question: what changed, and can it be repeated deliberately?", styles["Subtle"]))
story.append(PageBreak())

add_title("The investor-ready opportunity", "The biggest lever is not finding demand. It is protecting the value already being created.")
opp = [
    ["Opportunity", "Evidence", "Move"],
    ["Reduce operational leakage", f"{money(friction_value)} tied to returned/cancelled orders", "Audit descriptions, delivery promise, and customer support loops"],
    ["Scale high-quality channels", f"{top_source['ReferralSource']} leads at {money(top_source['Revenue'])}", "Shift test budget toward high-AOV sources"],
    ["Retain large baskets", f"{len(outliers)} outlier orders above {money(upper_fence)}", "Create a VIP follow-up workflow"],
    ["Broaden winning assortment", f"{top_product['Product']} leads but revenue is spread", "Package complementary products into bundles"],
]
tbl = PdfTable(opp, colWidths=[2.4*inch, 3.3*inch, 4.4*inch])
tbl.setStyle(TableStyle([
    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#" + NAVY)),
    ("TEXTCOLOR", (0,0), (-1,0), colors.white),
    ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ("BACKGROUND", (0,1), (-1,-1), colors.HexColor("#" + CREAM)),
    ("TEXTCOLOR", (0,1), (-1,-1), colors.HexColor("#" + NAVY)),
    ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#" + LINE)),
    ("VALIGN", (0,0), (-1,-1), "TOP"),
    ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#" + CREAM), colors.white]),
]))
story.append(tbl)
story.append(Spacer(1, 0.25 * inch))
story.append(Paragraph("Bottom line: this dataset can support a credible early ecommerce story, provided the next phase validates why cancellations and returns are happening and proves that high-value order behavior can be repeated.", styles["Big"]))

doc.build(story)

summary = {
    "total_order_value": total_value,
    "orders": total_orders,
    "average_order_value": avg_order,
    "returned_cancelled_value": friction_value,
    "outlier_count": int(len(outliers)),
    "top_product": top_product.to_dict(),
    "top_source": top_source.to_dict(),
    "best_month": best_month.to_dict(),
    "latest_month_mom_pct": float(latest_mom),
}
(OUTDIR / "eda_summary.json").write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")

print(xlsx_path)
print(pdf_path)
